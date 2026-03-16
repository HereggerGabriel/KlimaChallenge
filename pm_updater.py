"""
pm_updater.py — KlimaChallenge PM Excel Updater
================================================
Designed for the v4 seven-sheet structure. Claude Code runs this at the
end of each session to update all structured data. Narrative columns
(Reflection, Theory Reference, detailed diary text) are left with
placeholder prompts so the human can enrich them via the web interface.

Usage: fill in the __main__ block at the bottom and run:
    python pm_updater.py

Sheets handled:
    📊 Dashboard       — KPI numbers + velocity table
    📓 Session Diary   — one row per session (structured fields only)
    📐 Decision Log    — ADR entries when a major decision was made
    📋 Backlog         — mark items done, add new items
    ⚠️ Gotchas         — add newly discovered platform rules
    🎯 Milestones      — update milestone status / completion
    ✅ Feature Registry — add newly shipped features
"""

import sys
sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Palette (must match build_pm_v3.py exactly) ───────────────────────────────
BLUE_DARK    = "00334F"
BLUE_MID     = "0B6C8A"
BLUE_XLIGHT  = "D6EEF5"
GREEN_DARK   = "418880"
GREEN_MID    = "3FB28F"
GREEN_XLIGHT = "E8F5E9"
RED_MID      = "E95B20"
RED_XLIGHT   = "FFF0ED"
AMBER        = "F59E0B"
AMBER_LIGHT  = "FEF3C7"
WHITE        = "FFFFFF"
OFF_WHITE    = "F8FAFB"
LIGHT_GRAY   = "EDF2F5"
MID_GRAY     = "C8D8E4"
TEXT_DARK    = "1A2E3B"
TEXT_MID     = "2D4A5A"
TEXT_LIGHT   = "5A7A8A"
PURPLE       = "7C3AED"
PURPLE_LIGHT = "EDE9FE"

PLACEHOLDER  = "[To be completed via web interface]"

# ── Style helpers ─────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=9, color=TEXT_DARK, italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Arial")

def _align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(color=MID_GRAY):
    s = Side(style="thin", color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def _style(cell, bg=WHITE, fg=TEXT_DARK, bold=False, size=9,
           h="left", v="center", wrap=True, italic=False, border_color=MID_GRAY):
    """Apply full style to a single cell."""
    cell.fill    = _fill(bg)
    cell.font    = _font(bold=bold, size=size, color=fg, italic=italic)
    cell.alignment = _align(h=h, v=v, wrap=wrap)
    cell.border  = _border(color=border_color)

def _row_bg(row_index, shade_even=True):
    """Alternating row background — even rows get LIGHT_GRAY."""
    return LIGHT_GRAY if (row_index % 2 == 0) == shade_even else WHITE


# ─────────────────────────────────────────────────────────────────────────────
class PMUpdater:
    """
    Loads the v4 KlimaChallenge PM Excel file and exposes methods for
    every routine post-session update. All methods are idempotent:
    they skip silently if the target row already exists.

    Intended workflow
    -----------------
    1. Claude Code fills in the __main__ block below.
    2. Runs this script at the end of each session.
    3. Structured data (KPIs, new rows, status changes) are updated.
    4. Narrative columns are populated with PLACEHOLDER text.
    5. Human enriches narrative columns via the web interface.
    """

    def __init__(self, path: str):
        if not os.path.exists(path):
            raise FileNotFoundError(f"PM file not found: {path}")
        self.path = path
        self.wb   = load_workbook(path)
        # Sheet name → workbook sheet mapping (tolerates emoji prefixes)
        self._sheets = {ws.title: ws for ws in self.wb.worksheets}

    def save(self, path: str = None):
        """Save to the original path or a new path if specified."""
        out = path or self.path
        self.wb.save(out)
        print(f"  ✓ Saved: {out}")

    def _get_sheet(self, name_fragment: str):
        """Find a sheet by partial name match (handles emoji prefixes)."""
        for title, ws in self._sheets.items():
            if name_fragment.lower() in title.lower():
                return ws
        raise KeyError(f"Sheet containing '{name_fragment}' not found. "
                       f"Available: {list(self._sheets.keys())}")

    def _find_last_data_row(self, ws, start_row: int, col: str = "A") -> int:
        """
        Walk down from start_row in column col and return the last row
        that has a non-None value. Returns start_row - 1 if nothing found
        (i.e. the table is empty).
        """
        last = start_row - 1
        for row in ws.iter_rows(min_row=start_row, min_col=1, max_col=1):
            cell = row[0]
            if cell.value is not None:
                last = cell.row
        return last

    def _find_row_containing(self, ws, search_value, col_index: int = 1,
                              start_row: int = 1) -> int:
        """
        Return the row number of the first cell in col_index that contains
        search_value (case-insensitive substring match). Returns -1 if not found.
        """
        for row in ws.iter_rows(min_row=start_row):
            cell = row[col_index - 1]
            if cell.value and search_value.lower() in str(cell.value).lower():
                return cell.row
        return -1

    # ──────────────────────────────────────────────────────────────────────────
    # 📊 DASHBOARD
    # ──────────────────────────────────────────────────────────────────────────

    def update_dashboard_kpis(self, sessions: int, features: int,
                               backlog: int, tech_debt: int, blockers: int):
        """
        Shift current KPI values (row 6) into the 'previous' note (row 7),
        then write new values into row 6.

        Parameters
        ----------
        sessions   : total sessions completed to date
        features   : total features shipped to date
        backlog    : open backlog items remaining AFTER this session
        tech_debt  : open tech debt items remaining AFTER this session
        blockers   : pre-production blockers remaining AFTER this session
        """
        ws = self._get_sheet("Dashboard")
        print(f"  Updating Dashboard KPIs — sessions={sessions}, "
              f"features={features}, backlog={backlog}, "
              f"tech_debt={tech_debt}, blockers={blockers}")

        # Read current row-6 values as 'previous' before overwriting
        prev_vals = {}
        for merged in ws.merged_cells.ranges:
            top_left = ws.cell(merged.min_row, merged.min_col)
            if merged.min_row == 6:
                prev_vals[merged.min_col] = top_left.value

        # The five KPI value cells are at columns A, C, E, G, I (row 6)
        # They are merged pairs: A6:B6, C6:D6, E6:F6, G6:H6, I6:J6
        new_vals = {1: sessions, 3: features, 5: backlog, 7: tech_debt, 9: blockers}
        color_map = {1: BLUE_DARK, 3: GREEN_DARK, 5: AMBER, 7: GREEN_MID, 9: RED_MID}

        for col, val in new_vals.items():
            cell = ws.cell(row=6, column=col)
            cell.value = str(val)
            _style(cell, bg=LIGHT_GRAY, fg=color_map[col],
                   bold=True, size=22, h="center", wrap=False)

        # Update the note row (row 7) with the new previous values
        prev_str = (f"prev sessions: {prev_vals.get(1,'–')} / "
                    f"{prev_vals.get(3,'–')} / {prev_vals.get(5,'–')} / "
                    f"{prev_vals.get(7,'–')} / {prev_vals.get(9,'–')}   "
                    "· Open Backlog = open feature/bug items; "
                    "Pre-Prod Blockers = separate infra items not counted above")
        note_cell = ws.cell(row=7, column=1)
        note_cell.value = prev_str
        _style(note_cell, bg=OFF_WHITE, fg=TEXT_LIGHT, italic=True, size=8)

    def add_velocity_row(self, session: str, date: str, focus: str,
                          features_n: int, category: str):
        """
        Append a new row to the velocity table on the Dashboard.
        The table starts with a header at row 21 and data from row 22 onward.

        Parameters
        ----------
        session    : e.g. "S13"
        date       : e.g. "Mar 17"
        focus      : short description of the session's main theme
        features_n : number of features shipped (0 for debt/infra sessions)
        category   : e.g. "Feature / Bug Fix"
        """
        ws = self._get_sheet("Dashboard")

        # Find last velocity data row (header is row 21, data starts row 22)
        last_r = self._find_last_data_row(ws, start_row=22)

        # Idempotency: skip if this session already exists in the table
        for row in ws.iter_rows(min_row=22, max_col=1):
            if row[0].value == session:
                print(f"  ⚠️  Velocity row for {session} already exists — skipping.")
                return

        r = last_r + 1
        ws.row_dimensions[r].height = 18

        # Compute cumulative total by summing column D from row 22 to last_r
        cumulative = 0
        for row in ws.iter_rows(min_row=22, max_row=last_r, min_col=4, max_col=4):
            for cell in row:
                if isinstance(cell.value, int):
                    cumulative += cell.value
        cumulative += features_n

        shade = (r % 2 == 0)
        bg = LIGHT_GRAY if shade else WHITE
        max_feats = 9
        bar_n = round((features_n / max_feats) * 10) if features_n > 0 else 0
        bar = "█" * bar_n + "░" * (10 - bar_n)
        bar_color = GREEN_MID if features_n >= 5 else (
            GREEN_DARK if features_n > 0 else TEXT_LIGHT)

        cols = [
            ("A", session,    {"bold": True, "fg": BLUE_MID, "wrap": False}),
            ("B", date,       {}),
            ("C", focus,      {"wrap": True}),
            ("D", features_n if features_n > 0 else "—",
                  {"h": "center",
                   "fg": GREEN_DARK if features_n > 0 else TEXT_LIGHT,
                   "bold": features_n >= 5}),
            ("E", category,   {"fg": TEXT_MID, "italic": True}),
            ("F", cumulative if features_n > 0 else "—",
                  {"h": "center", "fg": BLUE_MID, "bold": True}),
            ("G", bar,        {"fg": bar_color}),
        ]
        for col_letter, val, style in cols:
            cell = ws[f"{col_letter}{r}"]
            cell.value = val
            _style(cell, bg=bg,
                   fg=style.get("fg", TEXT_DARK),
                   bold=style.get("bold", False),
                   italic=style.get("italic", False),
                   h=style.get("h", "left"),
                   wrap=style.get("wrap", False))

        print(f"  Added velocity row: {session} · {features_n} features · "
              f"cumulative {cumulative}")

    # ──────────────────────────────────────────────────────────────────────────
    # 📓 SESSION DIARY
    # ──────────────────────────────────────────────────────────────────────────

    def add_session_row(self, session: str, date: str, scope: str,
                         what: str, why: str, problems: str,
                         alternatives: str, decisions: str,
                         features_n: int, effort: str, thesis_tag: str,
                         reflection: str = PLACEHOLDER):
        """
        Insert a new session row into the Session Diary, just before
        the colour-legend row at the bottom of the table.

        The `reflection` parameter defaults to a placeholder — it is
        intentionally left for enrichment via the web interface after
        the session, when retrospective distance is possible.

        Parameters
        ----------
        session      : e.g. "S13"
        date         : e.g. "2026-03-17"
        scope        : one-line session theme
        what         : deliverables and files changed
        why          : motivation and design rationale
        problems     : problems encountered
        alternatives : approaches rejected
        decisions    : standing rules / gotchas for future reference
        features_n   : number of features shipped
        effort       : estimated effort e.g. "~4h"
        thesis_tag   : e.g. "Gamification Theory / UX Design"
        reflection   : retrospective reflection (leave as PLACEHOLDER)
        """
        ws = self._get_sheet("Session Diary")

        # Idempotency: skip if session already exists
        if self._find_row_containing(ws, session, col_index=1, start_row=5) != -1:
            print(f"  ⚠️  Session Diary row for {session} already exists — skipping.")
            return

        # Find the legend row (contains "Column colour guide")
        # Insert before it so the legend stays at the bottom
        legend_r = self._find_row_containing(
            ws, "Column colour guide", col_index=1, start_row=5)

        if legend_r == -1:
            # No legend row found — append after last data row
            legend_r = self._find_last_data_row(ws, start_row=5) + 2

        # Insert a blank row at legend_r, pushing everything down
        ws.insert_rows(legend_r)
        r = legend_r

        # Count existing data rows to determine alternating shade
        data_rows = legend_r - 5  # rows 5..(legend_r-1) were data rows
        shade = (data_rows // 2) % 2 == 0
        bg = LIGHT_GRAY if shade else WHITE

        ws.row_dimensions[r].height = 90

        # Column order matches the v3 header: A-L
        row_data = [
            ("A", session,      {"bold": True, "fg": BLUE_MID, "v": "top",
                                 "h": "center", "wrap": False}),
            ("B", date,         {"fg": TEXT_MID, "v": "top"}),
            ("C", scope,        {"bold": True, "fg": BLUE_DARK, "v": "top"}),
            ("D", what,         {"v": "top"}),
            ("E", why,          {"fg": GREEN_DARK, "v": "top"}),
            ("F", problems,     {"fg": RED_MID, "v": "top"}),
            ("G", alternatives, {"fg": TEXT_MID, "italic": True, "v": "top"}),
            ("H", decisions,    {"fg": BLUE_MID, "v": "top"}),
            ("I", str(features_n) if features_n > 0 else "—",
                  {"h": "center", "v": "top",
                   "fg": GREEN_MID if features_n > 0 else TEXT_LIGHT,
                   "bold": features_n > 0}),
            ("J", effort,       {"h": "center", "v": "top", "fg": TEXT_MID}),
            ("K", thesis_tag,   {"fg": AMBER, "italic": True, "v": "top"}),
            ("L", reflection,   {"fg": PURPLE, "v": "top",
                                 "italic": reflection == PLACEHOLDER}),
        ]

        for col_letter, val, style in row_data:
            cell = ws[f"{col_letter}{r}"]
            cell.value = val
            _style(cell, bg=bg,
                   fg=style.get("fg", TEXT_DARK),
                   bold=style.get("bold", False),
                   italic=style.get("italic", False),
                   h=style.get("h", "left"),
                   v=style.get("v", "center"),
                   wrap=True)

        print(f"  Added Session Diary row: {session} · {scope} · "
              f"{features_n} features")
        if reflection == PLACEHOLDER:
            print(f"  ℹ️  Reflection column left as placeholder — "
                  f"enrich via web interface after the session.")

    # ──────────────────────────────────────────────────────────────────────────
    # 📐 DECISION LOG
    # ──────────────────────────────────────────────────────────────────────────

    def add_adr(self, n: int, area: str, context: str, options: str,
                 decision: str, criteria: str, reversal: str, session: str,
                 bg_color: str = BLUE_XLIGHT):
        """
        Append a new Architecture Decision Record to the Decision Log.

        Parameters
        ----------
        n          : ADR number (integer)
        area       : short name e.g. "Chart Library Choice"
        context    : what forced this decision
        options    : options that were considered (pipe-separated or prose)
        decision   : what was decided and why
        criteria   : criteria used to evaluate options
        reversal   : what would need to change if this decision is reversed
        session    : session in which the decision was made e.g. "S13"
        bg_color   : row background colour (default BLUE_XLIGHT)
        """
        ws = self._get_sheet("Decision Log")

        # Idempotency: skip if ADR number or area already exists
        for row in ws.iter_rows(min_row=5, max_col=2):
            if str(row[0].value) == str(n) or (
                    row[1].value and area.lower() in row[1].value.lower()):
                print(f"  ⚠️  ADR #{n} ({area}) already exists — skipping.")
                return

        last_r = self._find_last_data_row(ws, start_row=5)
        r = last_r + 1
        ws.row_dimensions[r].height = 80

        row_data = [
            ("A", str(n),    {"bold": True, "fg": BLUE_MID, "h": "center"}),
            ("B", area,      {"bold": True, "fg": BLUE_DARK}),
            ("C", context,   {"fg": TEXT_DARK}),
            ("D", options,   {"fg": TEXT_MID, "italic": True}),
            ("E", decision,  {"fg": GREEN_DARK}),
            ("F", criteria,  {"fg": TEXT_DARK}),
            ("G", reversal,  {"fg": RED_MID}),
            ("H", session,   {"h": "center", "fg": BLUE_MID, "bold": True}),
        ]

        for col_letter, val, style in row_data:
            cell = ws[f"{col_letter}{r}"]
            cell.value = val
            _style(cell, bg=bg_color,
                   fg=style.get("fg", TEXT_DARK),
                   bold=style.get("bold", False),
                   italic=style.get("italic", False),
                   h=style.get("h", "left"),
                   v="top")

        print(f"  Added ADR #{n}: {area}")

    # ──────────────────────────────────────────────────────────────────────────
    # 📋 BACKLOG
    # ──────────────────────────────────────────────────────────────────────────

    def mark_backlog_done(self, identifier: str, session_completed: str,
                           notes: str = ""):
        """
        Find a backlog item by its # number or task name substring and
        mark it as done. Updates the Status cell to "✅ Done", changes
        the row fill to GREEN_XLIGHT, and appends to the Notes column.

        Parameters
        ----------
        identifier        : item # as string (e.g. "36") or task name fragment
        session_completed : e.g. "S13"
        notes             : optional completion note appended to existing notes
        """
        ws = self._get_sheet("Backlog")

        # Search column A (item #) then column C (task name)
        target_row = -1
        for row in ws.iter_rows(min_row=6):
            num_cell  = row[0]   # column A
            task_cell = row[2]   # column C
            if (str(num_cell.value) == str(identifier) or
                    (task_cell.value and
                     identifier.lower() in str(task_cell.value).lower())):
                target_row = num_cell.row
                break

        if target_row == -1:
            print(f"  ⚠️  Backlog item '{identifier}' not found.")
            return

        ws.row_dimensions[target_row].height = 28

        # Update status column (G = index 7)
        status_cell = ws.cell(row=target_row, column=7)
        status_cell.value = "✅ Done"
        _style(status_cell, bg=GREEN_XLIGHT, fg=GREEN_DARK, bold=True)

        # Update target column (H = index 8) with session completed
        target_cell = ws.cell(row=target_row, column=8)
        target_cell.value = session_completed
        _style(target_cell, bg=GREEN_XLIGHT, fg=GREEN_DARK)

        # Append to notes column (I = index 9)
        notes_cell = ws.cell(row=target_row, column=9)
        existing   = notes_cell.value or ""
        suffix     = f"  ·  Completed {session_completed}"
        if notes:
            suffix += f" — {notes}"
        notes_cell.value = existing + suffix
        _style(notes_cell, bg=GREEN_XLIGHT, fg=TEXT_MID, italic=True)

        # Re-shade all other cells in the row to green
        for col_idx in range(1, 7):
            cell = ws.cell(row=target_row, column=col_idx)
            cell.fill = _fill(GREEN_XLIGHT)
            cell.font = _font(size=9, color=TEXT_LIGHT,
                              bold=(col_idx in [1, 3]))

        print(f"  Marked backlog item '{identifier}' as done ({session_completed})")

    def add_backlog_item(self, num: str, category: str, task: str,
                          description: str, priority: str, effort: str,
                          target: str, notes: str, status: str = "⬜ Todo"):
        """
        Add a new item to the OPEN section of the backlog, inserted just
        before the Pre-Production Checklist section header.

        Parameters
        ----------
        num         : item number as string, e.g. "41" or "—"
        category    : e.g. "Feature", "Bug Fix", "Tech Debt", "UX"
        task        : short task name
        description : full description of the work required
        priority    : e.g. "🔴 Critical", "🟠 High", "🟡 Medium", "🔵 Low"
        effort      : e.g. "XS (30m)", "S (2-4h)", "M (4-8h)", "L (1-2d)"
        target      : e.g. "S13", "Any", "Pre-launch"
        notes       : implementation notes, links to ADRs, etc.
        status      : defaults to "⬜ Todo"
        """
        ws = self._get_sheet("Backlog")

        # Idempotency: skip if item number or task already exists
        for row in ws.iter_rows(min_row=6, max_col=3):
            if (str(row[0].value) == str(num) or
                    (row[2].value and
                     task.lower() in str(row[2].value).lower())):
                print(f"  ⚠️  Backlog item '{num} — {task}' already exists — skipping.")
                return

        # Find the "PRE-PRODUCTION CHECKLIST" section header to insert before it
        pre_prod_r = self._find_row_containing(
            ws, "PRE-PRODUCTION", col_index=1, start_row=6)
        if pre_prod_r == -1:
            # Fall back to appending after the last open item
            pre_prod_r = self._find_last_data_row(ws, start_row=6) + 1

        ws.insert_rows(pre_prod_r)
        r = pre_prod_r

        # Unmerge any cells on this row (insert_rows can land inside a merged range)
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row <= r <= mr.max_row:
                ws.unmerge_cells(str(mr))

        priority_fills = {
            "🔴 Critical": RED_XLIGHT,
            "🟠 High":     "FFF4EC",
            "🟡 Medium":   AMBER_LIGHT,
            "🔵 Low":      BLUE_XLIGHT,
        }
        priority_colors = {
            "🔴 Critical": RED_MID,
            "🟠 High":     "E95B20",
            "🟡 Medium":   AMBER,
            "🔵 Low":      BLUE_MID,
        }
        bg = priority_fills.get(priority, LIGHT_GRAY)

        ws.row_dimensions[r].height = 48
        for col_idx, (val, style) in enumerate([
            (str(num),      {"bold": True, "fg": TEXT_DARK}),
            (category,      {}),
            (task,          {"bold": True}),
            (description,   {}),
            (priority,      {"fg": priority_colors.get(priority, TEXT_DARK),
                             "bold": "Critical" in priority}),
            (effort,        {}),
            (status,        {}),
            (target,        {}),
            (notes,         {"italic": True, "fg": TEXT_MID}),
        ], start=1):
            cell = ws.cell(row=r, column=col_idx)
            cell.value = val
            _style(cell, bg=bg,
                   fg=style.get("fg", TEXT_DARK),
                   bold=style.get("bold", False),
                   italic=style.get("italic", False))

        print(f"  Added backlog item: {num} — {task} ({priority})")

    # ──────────────────────────────────────────────────────────────────────────
    # ⚠️ GOTCHAS
    # ──────────────────────────────────────────────────────────────────────────

    def add_gotcha(self, area: str, rule: str, why: str, broken: str,
                    session: str, severity: str, thesis: str):
        """
        Append a new gotcha to the Gotchas sheet.

        Parameters
        ----------
        area     : short area name e.g. "Supabase session refresh"
        rule     : the rule to follow in one clear sentence
        why      : full context — why does this rule exist?
        broken   : what breaks if the rule is ignored (be specific)
        session  : session where this was discovered e.g. "S13"
        severity : "🔴 Critical" | "🟠 High" | "🟡 Medium" | "🔵 Info"
        thesis   : e.g. "Platform limitations / Auth"
        """
        ws = self._get_sheet("Gotchas")

        # Idempotency: skip if area already exists
        if self._find_row_containing(ws, area, col_index=1, start_row=5) != -1:
            print(f"  ⚠️  Gotcha '{area}' already exists — skipping.")
            return

        sev_bg = {
            "🔴 Critical": RED_XLIGHT,
            "🟠 High":     "FFF4EC",
            "🟡 Medium":   AMBER_LIGHT,
            "🔵 Info":     BLUE_XLIGHT,
        }
        sev_fg = {
            "🔴 Critical": RED_MID,
            "🟠 High":     "E95B20",
            "🟡 Medium":   AMBER,
            "🔵 Info":     BLUE_MID,
        }

        last_r = self._find_last_data_row(ws, start_row=5)
        r = last_r + 1
        ws.row_dimensions[r].height = 56
        bg = sev_bg.get(severity, WHITE)

        for col_letter, val, style in [
            ("A", area,     {"bold": True, "fg": BLUE_DARK}),
            ("B", rule,     {"bold": True,
                             "fg": RED_MID if "Critical" in severity
                             else TEXT_DARK}),
            ("C", why,      {}),
            ("D", broken,   {"fg": RED_MID, "italic": True}),
            ("E", session,  {"h": "center", "fg": BLUE_MID}),
            ("F", severity, {"h": "center",
                             "fg": sev_fg.get(severity, TEXT_DARK),
                             "bold": True}),
            ("G", thesis,   {"fg": AMBER, "italic": True}),
        ]:
            cell = ws[f"{col_letter}{r}"]
            cell.value = val
            _style(cell, bg=bg,
                   fg=style.get("fg", TEXT_DARK),
                   bold=style.get("bold", False),
                   italic=style.get("italic", False),
                   h=style.get("h", "left"))

        print(f"  Added gotcha: {area} ({severity})")

    # ──────────────────────────────────────────────────────────────────────────
    # 🎯 MILESTONES
    # ──────────────────────────────────────────────────────────────────────────

    def update_milestone(self, milestone_id: str, status: str,
                          completion: str, blockers: str = "—"):
        """
        Find a milestone row by its ID (e.g. "M7") and update its
        status, completion percentage, and blockers columns.

        Parameters
        ----------
        milestone_id : e.g. "M7"
        status       : e.g. "✅ Complete" or "🔄 In Progress"
        completion   : e.g. "100%" or "60%"
        blockers     : updated blockers string, or "—" if none
        """
        ws = self._get_sheet("Milestones")
        target_row = self._find_row_containing(
            ws, milestone_id, col_index=1, start_row=6)

        if target_row == -1:
            print(f"  ⚠️  Milestone '{milestone_id}' not found.")
            return

        done = status.startswith("✅")
        bg   = GREEN_XLIGHT if done else AMBER_LIGHT
        color = GREEN_DARK if done else AMBER

        # Column D = status (col 4), F = blockers (col 6)
        # Re-shade all cells in the row
        for col_idx in range(1, 8):
            cell = ws.cell(row=target_row, column=col_idx)
            cell.fill = _fill(bg)

        status_cell = ws.cell(row=target_row, column=4)
        status_cell.value = f"{status}  ({completion})"
        _style(status_cell, bg=bg, fg=color, bold=True)

        blockers_cell = ws.cell(row=target_row, column=6)
        blockers_cell.value = blockers
        _style(blockers_cell, bg=bg, fg=TEXT_DARK)

        print(f"  Updated milestone {milestone_id}: {status} ({completion})")

    # ──────────────────────────────────────────────────────────────────────────
    # ✅ FEATURE REGISTRY
    # ──────────────────────────────────────────────────────────────────────────

    def add_feature(self, num: int, name: str, category: str,
                     session: str, files: str, description: str,
                     xp_impact: str, also_modified: str = "—",
                     theory_ref: str = PLACEHOLDER,
                     thesis_tag: str = "—"):
        """
        Append a new feature row to the Feature Registry.

        The `theory_ref` parameter defaults to PLACEHOLDER — for
        gamification features this should be enriched via the web
        interface with appropriate academic citations.

        Parameters
        ----------
        num           : feature number (integer)
        name          : feature name
        category      : e.g. "Feature", "UI", "UX", "Gamification", "Bug Fix"
        session       : session delivered e.g. "S13"
        files         : primary files affected
        description   : one-sentence description
        xp_impact     : e.g. "—", "Yes", "Affects XP", "Display"
        also_modified : sessions where the feature was later modified
        theory_ref    : academic theory reference (leave as PLACEHOLDER
                        if not yet determined)
        thesis_tag    : e.g. "Gamification Theory"
        """
        ws = self._get_sheet("Feature Registry")

        # Idempotency: skip if feature number already exists
        for row in ws.iter_rows(min_row=5, max_col=1):
            if str(row[0].value) == str(num):
                print(f"  ⚠️  Feature #{num} ({name}) already exists — skipping.")
                return

        cat_bg = {
            "Core":         BLUE_XLIGHT,
            "UI":           PURPLE_LIGHT,
            "UX":           GREEN_XLIGHT,
            "Auth":         "FEF9C3",
            "Onboarding":   "FEF9C3",
            "Feature":      "E0F2FE",
            "Gamification": GREEN_XLIGHT,
            "Settings":     LIGHT_GRAY,
            "Refactor":     LIGHT_GRAY,
            "Bug Fix":      RED_XLIGHT,
            "Stats":        BLUE_XLIGHT,
            "PM":           LIGHT_GRAY,
        }
        thesis_map = {
            "Core":         "Data Architecture",
            "UI":           "Interface Design",
            "UX":           "UX Design",
            "Auth":         "Auth Architecture",
            "Onboarding":   "FRE Design",
            "Feature":      "Feature Design",
            "Gamification": "Gamification Theory",
            "Settings":     "User Configuration",
            "Refactor":     "Software Quality",
            "Bug Fix":      "Bug Analysis",
            "Stats":        "Data Visualisation",
            "PM":           "Development Process",
        }

        bg = cat_bg.get(category, WHITE)
        if thesis_tag == "—":
            thesis_tag = thesis_map.get(category, "—")

        last_r = self._find_last_data_row(ws, start_row=5)
        r = last_r + 1
        ws.row_dimensions[r].height = 30

        row_data = [
            ("A", num,          {"h": "center", "bold": True, "fg": BLUE_MID}),
            ("B", name,         {"bold": True}),
            ("C", category,     {"fg": TEXT_MID}),
            ("D", session,      {"h": "center", "fg": BLUE_MID}),
            ("E", also_modified,{"fg": TEXT_MID, "italic": True, "h": "center"}),
            ("F", files,        {"fg": TEXT_MID, "italic": True}),
            ("G", description,  {}),
            ("H", xp_impact,    {"h": "center",
                                  "fg": GREEN_DARK if xp_impact not in
                                  ["—", "Display"] else TEXT_LIGHT}),
            ("I", theory_ref,   {"fg": PURPLE,
                                  "italic": theory_ref == PLACEHOLDER}),
            ("J", thesis_tag,   {"fg": AMBER, "italic": True}),
        ]

        for col_letter, val, style in row_data:
            cell = ws[f"{col_letter}{r}"]
            cell.value = str(val) if not isinstance(val, str) else val
            _style(cell, bg=bg,
                   fg=style.get("fg", TEXT_DARK),
                   bold=style.get("bold", False),
                   italic=style.get("italic", False),
                   h=style.get("h", "left"))

        print(f"  Added feature #{num}: {name} ({category}, {session})")
        if theory_ref == PLACEHOLDER:
            print(f"  ℹ️  Theory Reference left as placeholder — "
                  f"enrich via web interface if gamification-related.")

    def update_feature_modified(self, num: int, session: str):
        """
        Add a session to the 'Also Modified In' column (E) of an
        existing feature row.

        Parameters
        ----------
        num     : feature number
        session : session to append e.g. "S13"
        """
        ws = self._get_sheet("Feature Registry")
        for row in ws.iter_rows(min_row=5, max_col=5):
            if str(row[0].value) == str(num):
                cell = row[4]  # column E = index 4
                existing = cell.value or "—"
                if session in existing:
                    print(f"  ⚠️  Feature #{num} already shows {session} in "
                          f"'Also Modified In' — skipping.")
                    return
                cell.value = existing.replace("—", "").strip()
                cell.value = (cell.value + f", {session}").lstrip(", ")
                print(f"  Updated feature #{num} 'Also Modified In' → {cell.value}")
                return
        print(f"  ⚠️  Feature #{num} not found in Feature Registry.")


# ─────────────────────────────────────────────────────────────────────────────
# __main__ — Claude Code fills this block and runs the script each session
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":

    # ── PATH ─────────────────────────────────────────────────────────────────
    PM_PATH = r"f:\projects\TravelApp\travelapp\KlimaChallenge_PM_v4.xlsx"

    pm = PMUpdater(PM_PATH)

    print("\n=== KlimaChallenge PM Updater ===\n")

    # ── 1. DASHBOARD KPIs ─────────────────────────────────────────────────────
    # Update with the new totals AFTER this session.
    # sessions   = total sessions completed including this one
    # features   = cumulative features shipped (all sessions)
    # backlog    = open backlog items remaining AFTER this session
    # tech_debt  = open tech debt items remaining AFTER this session
    # blockers   = pre-production blockers remaining AFTER this session
    pm.update_dashboard_kpis(
        sessions  = 13,
        features  = 39,
        backlog   = 14,   # #41 now done; 11 new items added; net open = 14
        tech_debt = 0,
        blockers  = 3,
    )

    # ── 2. VELOCITY ROW ───────────────────────────────────────────────────────
    pm.add_velocity_row(
        session    = "S13",
        date       = "Mar 17",
        focus      = "Bug fixes #36/#37/#41 + UX #38 + backlog brainstorm (11 new items)",
        features_n = 4,
        category   = "Bug Fix / UX",
    )

    # ── 3. SESSION DIARY ROW ─────────────────────────────────────────────────
    pm.add_session_row(
        session      = "S13",
        date         = "2026-03-17",
        scope        = "Bug fixes #36/#37/#41 + UX fix #38 + full backlog brainstorm",
        what         = (
            "Fixed #36: handleDeleteAllTrips now clears @claimedQuests, @claimedAchievements, "
            "@dailyQuestSelection, @weeklyQuestSelection via AsyncStorage.multiRemove. "
            "Fixed #37: loadTrips() fallback changed from initialTrips to []. "
            "Fixed #38: KlimaTicket dropdown closes on outside tap via absolute TouchableWithoutFeedback overlay. "
            "Fixed #41: Stats screen — proper empty state (icon + CTA) replaces broken zero-data sections when totalTrips=0. "
            "Quests screen — info banner added when trips.length=0. "
            "Backlog brainstorm: 11 new items #41-#51 added. "
            "pm_updater.py replaced with new v4-compatible version; merged cell bug in add_backlog_item fixed. "
            "Files: app/stats.tsx, app/quests.tsx, app/profile.tsx, services/tripStorage.ts, pm_updater.py"
        ),
        why          = (
            "#36/#37 are pre-prod blockers. "
            "#38: open dropdown with no dismiss was a jarring UX gap. "
            "#41: new users (post #37 fix) hit blank/broken screens immediately — empty states are critical for first impression."
        ),
        problems     = (
            "pm_updater merged cell bug: add_backlog_item landed on openpyxl merged cell rows, "
            "silently writing only column A. Fixed by calling ws.unmerge_cells() before writing."
        ),
        alternatives = (
            "#38: TouchableWithoutFeedback wrapping ScrollView considered (interferes with scroll) — "
            "absolute overlay chosen instead. "
            "#41 Stats: considered keeping KlimaTicket card in empty state — removed to keep empty state clean."
        ),
        decisions    = (
            "AsyncStorage.multiRemove is correct for atomic batch key deletion. "
            "Absolute overlay (zIndex 10) is the standard dismiss pattern for inline dropdowns inside ScrollViews. "
            "Stats empty state replaces ALL data sections when 0 trips — avoids zero-value noise."
        ),
        features_n   = 4,
        effort       = "~3h",
        thesis_tag   = "Bug Analysis / UX Design",
    )

    # ── 4. MARK BACKLOG ITEMS DONE ────────────────────────────────────────────
    pm.mark_backlog_done("36", "S13", "AsyncStorage.multiRemove for all 4 quest/achievement keys in handleDeleteAllTrips")
    pm.mark_backlog_done("37", "S13", "Changed fallback from `return initialTrips` to `return []` in tripStorage.ts")
    pm.mark_backlog_done("38", "S13", "Absolute TouchableWithoutFeedback overlay (zIndex 10) closes dropdown; dropdown at zIndex 11")
    pm.mark_backlog_done("41", "S13", "Stats: full empty state replacing data sections. Quests: info banner.")

    # ── 5. FEATURE REGISTRY ───────────────────────────────────────────────────
    pm.add_feature(
        num           = 36,
        name          = "Delete All — full quest/achievement state reset",
        category      = "Bug Fix",
        session       = "S13",
        files         = "app/profile.tsx",
        description   = "handleDeleteAllTrips now clears all 4 quest/achievement AsyncStorage keys atomically alongside trips+XP.",
        xp_impact     = "Yes",
    )
    pm.add_feature(
        num           = 37,
        name          = "Remove initialTrips demo data fallback",
        category      = "Bug Fix",
        session       = "S13",
        files         = "services/tripStorage.ts",
        description   = "loadTrips() now returns [] when storage key is absent — new users start with an empty trip list.",
        xp_impact     = "—",
    )
    pm.add_feature(
        num           = 38,
        name          = "Dropdown close on outside tap",
        category      = "UX",
        session       = "S13",
        files         = "app/profile.tsx",
        description   = "KlimaTicket dropdown dismisses when tapping outside via absolute overlay — consistent with modal dismiss patterns.",
        xp_impact     = "—",
    )
    pm.add_feature(
        num           = 39,
        name          = "Empty states — Stats & Quests screens",
        category      = "UX",
        session       = "S13",
        files         = "app/stats.tsx, app/quests.tsx",
        description   = "Stats: full empty state with icon + CTA replaces zero-data sections. Quests: info banner when trips=0.",
        xp_impact     = "—",
    )

    # ── 6. NEW BACKLOG ITEMS (brainstorm S13) ────────────────────────────────
    # 🔴 Critical
    pm.add_backlog_item(
        num         = "41",
        category    = "UX",
        task        = "Empty states — Stats & Quests screens",
        description = "New users (post #37 fix) land on Stats with empty/broken charts and Quests with all progress at 0. Both screens need proper empty state UI guiding users to add their first trip.",
        priority    = "🔴 Critical",
        effort      = "S (2-4h)",
        target      = "S14",
        notes       = "Stats: replace chart with illustration + CTA. Quests: show locked state with 'Add a trip to start earning XP' prompt.",
    )
    # 🟠 High
    pm.add_backlog_item(
        num         = "42",
        category    = "UX",
        task        = "Manual date/time for trips",
        description = "Users can only log trips at the current moment. No way to backfill yesterday's commute. A date/time picker on the Add Trip modal is needed for real-world usage.",
        priority    = "🟠 High",
        effort      = "S (2-4h)",
        target      = "Any",
        notes       = "Add DateTimePicker (expo-datetime-picker or RN DateTimePicker) to QuickAddTripModal. Default to now, allow past dates.",
    )
    pm.add_backlog_item(
        num         = "43",
        category    = "Gamification",
        task        = "Streak system",
        description = "Track consecutive days the user logs at least one trip. Display streak on UserLevelCard. Streak loss is a powerful retention mechanic (loss aversion). Fits thesis well.",
        priority    = "🟠 High",
        effort      = "M (4-8h)",
        target      = "Any",
        notes       = "AsyncStorage key @currentStreak + @lastTripDate. Streak breaks if no trip logged by midnight. Consider streak freeze item as future extension.",
    )
    pm.add_backlog_item(
        num         = "44",
        category    = "Onboarding",
        task        = "Onboarding rework package",
        description = (
            "Full onboarding overhaul before production: (1) new slide content + real images, "
            "(2) personalization step — set name + KlimaTicket type inline so users skip empty profile, "
            "(3) 'Add your first trip' prompt on completion. Package with PP3 (onboarding images)."
        ),
        priority    = "🟠 High",
        effort      = "L (1-2d)",
        target      = "Pre-launch",
        notes       = "Supersedes PP3. Low current priority but must ship before production. Personalization step reduces time-to-value significantly.",
    )
    # 🟡 Medium
    pm.add_backlog_item(
        num         = "45",
        category    = "UX",
        task        = "Haptic feedback",
        description = "Add haptic feedback on: trip add confirmation, XP toast, quest claim, level-up, slide-to-delete confirm. Quick win for perceived quality.",
        priority    = "🟡 Medium",
        effort      = "XS (1h)",
        target      = "Any",
        notes       = "Use expo-haptics. Haptics.impactAsync(ImpactFeedbackStyle.Medium) for most; Heavy for level-up.",
    )
    pm.add_backlog_item(
        num         = "46",
        category    = "UX",
        task        = "Trip list search and filter",
        description = "Once users have 30+ trips the flat list becomes unwieldy. Add search by origin/destination and filter by transport type and/or date range.",
        priority    = "🟡 Medium",
        effort      = "M (4-8h)",
        target      = "Any",
        notes       = "Filter bar above trip list in user.tsx. Client-side filtering on trips array — no API needed.",
    )
    # 🔵 Low
    pm.add_backlog_item(
        num         = "47",
        category    = "Feature",
        task        = "Recurring trip / one-tap repeat",
        description = "Allow users to repeat a previous trip with one tap — pre-fills all fields from the last matching trip. Reduces friction for daily commuters.",
        priority    = "🔵 Low",
        effort      = "S (2-4h)",
        target      = "Any",
        notes       = "Could surface as a 'Repeat' button in TripDetailModal or as a quick-action in the favorites section.",
    )
    pm.add_backlog_item(
        num         = "48",
        category    = "Feature",
        task        = "Push notifications",
        description = "Streak-at-risk reminder (evening if no trip logged), daily quest expiry warning, weekly quest reset notification.",
        priority    = "🔵 Low",
        effort      = "M (4-8h)",
        target      = "Any",
        notes       = "Requires expo-notifications + user permission prompt. Schedule local notifications — no backend needed for streak/quest reminders.",
    )
    pm.add_backlog_item(
        num         = "49",
        category    = "Feature",
        task        = "Shareable achievement cards",
        description = "When a user unlocks an achievement, offer a screenshot-friendly share card (achievement name, icon, app branding). Social loop and organic marketing.",
        priority    = "🔵 Low",
        effort      = "S (2-4h)",
        target      = "Any",
        notes       = "Use expo-sharing + react-native-view-shot to capture the card as an image. Trigger from achievement claim.",
    )
    pm.add_backlog_item(
        num         = "50",
        category    = "Tech Debt",
        task        = "Clean up stale XP seeding comment in user.tsx",
        description = "Comment in user.tsx references 'ensuring 10 hardcoded dummy trips grant XP on fresh installs' — no longer true after #37 fix. Misleading for future devs.",
        priority    = "🔵 Low",
        effort      = "XS (5m)",
        target      = "Any",
        notes       = "One-line comment update in the XP init useEffect block.",
    )
    pm.add_backlog_item(
        num         = "51",
        category    = "UX",
        task        = "Offline / API error indicator",
        description = "When ÖBB API is unreachable, connection search fails silently. Show a small inline 'Offline — manual entry only' notice in QuickAddTripModal when API calls fail.",
        priority    = "🔵 Low",
        effort      = "XS (1h)",
        target      = "Any",
        notes       = "Catch fetch errors in oebbApi.ts and surface via a state flag in QuickAddTripModal. No retry logic needed.",
    )

    # ═══════════════════════════════════════════════════════════════════════════
    # SESSION 14 — 2026-03-16
    # DateTimePicker + Recent Trips screen rework + Profile nav rework
    # ═══════════════════════════════════════════════════════════════════════════

    # ── S14-1. DASHBOARD KPIs ─────────────────────────────────────────────────
    pm.update_dashboard_kpis(
        sessions  = 14,
        features  = 42,   # 39 (S13) + 3 new (DateTimePicker, Trips screen, Profile nav)
        backlog   = 13,   # #42 now done; 13 open items remain
        tech_debt = 0,
        blockers  = 5,    # PP1-PP5 unchanged
    )

    # ── S14-2. VELOCITY ROW ───────────────────────────────────────────────────
    pm.add_velocity_row(
        session    = "S14",
        date       = "Mar 16",
        focus      = "DateTimePicker + full Trips screen + profile nav rework",
        features_n = 3,
        category   = "Feature / UX",
    )

    # ── S14-3. SESSION DIARY ROW ──────────────────────────────────────────────
    pm.add_session_row(
        session      = "S14",
        date         = "2026-03-16",
        scope        = "Native DateTimePicker · Trips list screen · Profile layout rework",
        what         = (
            "#42 DateTimePicker: replaced custom 7-day date scroll + manual HH:MM inputs with "
            "@react-native-community/datetimepicker in both QuickAddTripModal and TripDetailModal. "
            "iOS: spinner display inline + Done button. Android: system dialog. Single Date state (no hour/minute strings). "
            "Plugin added to app.config.ts. Requires custom dev build (not Expo Go). "
            "Recent Trips rework: main screen now groups preview trips by date (date group headers). "
            "New app/trips.tsx screen: filter chips (All/Bus/Train/Tram/Subway), full date-grouped list, "
            "swipe-left-to-delete (RNGH Pan, -80px threshold, red zone, 200ms fly-off). "
            "'See all X trips' navigates to /trips. New utils/tripGrouping.ts shared helper "
            "(formatDateLabel, groupTripsByDate). Deleted dead stubs app/new-trip.tsx + app/trip/[id].tsx. "
            "Profile rework: new layout order — Avatar → Display Name + KlimaTicket + Save → "
            "navSection (Stats / Achievements / My Trips) → Danger Zone. "
            "My Trips button added (MaterialIcons format-list-bulleted, routes to /trips). "
            "Files: app.config.ts, components/ui/QuickAddTripModal.tsx, components/ui/TripDetailModal.tsx, "
            "app/(tabs)/user.tsx, app/(tabs)/_user_styles.tsx, app/trips.tsx (new), "
            "app/profile.tsx, utils/tripGrouping.ts (new)"
        ),
        why          = (
            "DateTimePicker: users had no way to backfill past trips — critical real-world usage gap. "
            "Trips screen: inline 'show all' expansion had no filter/delete; a dedicated screen matches "
            "the mental model of a full trip history. "
            "Profile rework: display name + KlimaTicket were buried below nav shortcuts; "
            "moving them above improves discoverability and data-entry flow."
        ),
        problems     = (
            "CRLF + 4-tab indentation in user.tsx and profile.tsx caused Edit tool to repeatedly fail "
            "string matching. Diagnosed with cat -A, resolved by writing complete file rewrites. "
            "Typo in user.tsx Write (missing } in JSX) caused cascade of TS parse errors — fixed with targeted edit. "
            "theme='dark' prop removal from TripRouteFields caused TS error (prop removed in S11 tech debt)."
        ),
        alternatives = (
            "DateTimePicker: expo-datetime-picker considered but requires native module build anyway; "
            "@react-native-community/datetimepicker chosen as more stable and widely used. "
            "Trips screen: ScrollView-based swipe vs RNGH Pan — RNGH chosen for correct gesture priority "
            "negotiation with parent ScrollView (activeOffsetX + failOffsetY)."
        ),
        decisions    = (
            "DateTimePicker requires custom dev build — document in project state. "
            "Swipe card needs opaque backgroundColor (blue.dark) to hide semi-transparent card bleed over red delete zone. "
            "Cross-screen sync: trips.tsx writes directly to AsyncStorage; user.tsx picks up via useFocusEffect (no shared state needed). "
            "CRLF files must be fully rewritten — Edit tool cannot reliably match Windows line endings."
        ),
        features_n   = 3,
        effort       = "~5h",
        thesis_tag   = "UX Design / Interaction Design",
    )

    # ── S14-4. MARK BACKLOG ITEMS DONE ───────────────────────────────────────
    pm.mark_backlog_done(
        "42", "S14",
        "Native DateTimePicker in both modals; iOS spinner+Done, Android dialog; single Date state"
    )

    # ── S14-5. FEATURE REGISTRY ───────────────────────────────────────────────
    pm.add_feature(
        num         = 40,
        name        = "Native DateTimePicker for trip date/time",
        category    = "UX",
        session     = "S14",
        files       = "QuickAddTripModal.tsx, TripDetailModal.tsx, app.config.ts",
        description = (
            "Replaced custom 7-day scroll + HH:MM text inputs with @react-native-community/datetimepicker. "
            "iOS: spinner inline + Done button. Android: system dialog. Single Date state."
        ),
        xp_impact   = "—",
        thesis_tag  = "UX Design",
    )
    pm.add_feature(
        num         = 41,
        name        = "All Trips screen with filter + swipe-to-delete",
        category    = "Feature",
        session     = "S14",
        files       = "app/trips.tsx (new), utils/tripGrouping.ts (new), app/(tabs)/user.tsx",
        description = (
            "New /trips stack screen: filter chips, date-grouped list, RNGH swipe-to-delete. "
            "Main screen preview grouped by date with 'See all' navigation."
        ),
        xp_impact   = "Yes",
        thesis_tag  = "UX Design",
    )
    pm.add_feature(
        num         = 42,
        name        = "Profile nav rework + My Trips shortcut",
        category    = "UI",
        session     = "S14",
        files       = "app/profile.tsx",
        description = (
            "New profile layout: Display Name + KlimaTicket above nav shortcuts. "
            "navSection group (Stats / Achievements / My Trips). My Trips routes to /trips."
        ),
        xp_impact   = "—",
        thesis_tag  = "Interface Design",
    )

    # ── S14-6. GOTCHA: CRLF files + Edit tool ─────────────────────────────────
    pm.add_gotcha(
        area     = "CRLF files + Edit tool string matching",
        rule     = "Files with Windows CRLF line endings must be fully rewritten — never use targeted Edit.",
        why      = (
            "The Edit tool matches exact byte strings. CRLF (\\r\\n) line endings cause every match to "
            "fail silently or partially, leading to cascading TS errors or phantom rewrites. "
            "Diagnosed via cat -A; the only reliable fix is a full Write of the affected file."
        ),
        broken   = "Edit silently fails, file unchanged; downstream TS errors if a partial write occurs.",
        session  = "S14",
        severity = "🟠 High",
        thesis   = "Platform limitations / DX",
    )

    # ═══════════════════════════════════════════════════════════════════════════
    # SESSION 15 — 2026-03-16
    # Haptic feedback + Streak system
    # ═══════════════════════════════════════════════════════════════════════════

    # ── S15-1. DASHBOARD KPIs ─────────────────────────────────────────────────
    pm.update_dashboard_kpis(
        sessions  = 15,
        features  = 44,   # 42 (S14) + 2 new (haptics, streak)
        backlog   = 11,   # closed #45 + #43; 11 open remain
        tech_debt = 0,
        blockers  = 5,    # PP1-PP5 unchanged
    )

    # ── S15-2. VELOCITY ROW ───────────────────────────────────────────────────
    pm.add_velocity_row(
        session    = "S15",
        date       = "Mar 16",
        focus      = "Haptic feedback throughout + streak system + streak badge on level card",
        features_n = 2,
        category   = "Gamification / Polish",
    )

    # ── S15-3. SESSION DIARY ROW ──────────────────────────────────────────────
    pm.add_session_row(
        session      = "S15",
        date         = "2026-03-16",
        scope        = "Haptic feedback · Streak system · UserLevelCard streak badge",
        what         = (
            "#45 Haptic feedback: expo-haptics wired throughout the app. "
            "ImpactFeedbackStyle.Medium on trip add (QuickAdd) and swipe-delete in trips.tsx. "
            "ImpactFeedbackStyle.Light on favorite card long-press add. "
            "NotificationFeedbackType.Success on level-up, quest/achievement/main-quest claim. "
            "#43 Streak system: new utils/streakSystem.ts with computeStreak(trips) — "
            "counts consecutive calendar days ending today or yesterday, returns 0 if streak broken. "
            "Displayed in UserLevelCard footer as 🔥 Xd badge (Palette.red.light, MaterialIcons local-fire-department). "
            "streak computed via useMemo([trips]) in user.tsx and passed as prop. Hidden when streak = 0. "
            "Files: utils/streakSystem.ts (new), components/ui/UserLevelCard.tsx, "
            "app/(tabs)/user.tsx (Write), app/trips.tsx, app/quests.tsx"
        ),
        why          = (
            "Haptics are table-stakes polish for a native app — they make every interaction feel intentional. "
            "Streak is a core retention mechanic; showing it on the level card keeps it visible "
            "without requiring a dedicated screen."
        ),
        problems     = (
            "expo-haptics was already installed (expo SDK default) — no install needed. "
            "user.tsx required full Write (CRLF file)."
        ),
        alternatives = (
            "Considered storing streak in AsyncStorage to avoid recomputing — rejected because "
            "computing from the trips array is O(n) and always accurate; no sync risk."
        ),
        decisions    = (
            "Streak starts from today if today has a trip, otherwise from yesterday — "
            "prevents streak breaking just because the user hasn't logged today's trip yet."
        ),
        features_n   = 2,
        effort       = "~2h",
        thesis_tag   = "Gamification / Engagement",
    )

    # ── S15-4. MARK BACKLOG ITEMS DONE ────────────────────────────────────────
    pm.mark_backlog_done("45", "S15", "expo-haptics wired to all key interactions")
    pm.mark_backlog_done("43", "S15", "computeStreak utility + 🔥 badge on UserLevelCard")

    # ── S15-5. FEATURE REGISTRY ───────────────────────────────────────────────
    pm.add_feature(
        num         = 43,
        name        = "Haptic feedback",
        category    = "Polish",
        session     = "S15",
        files       = "user.tsx, trips.tsx, quests.tsx",
        description = (
            "expo-haptics across all key interactions: trip add (Medium), favorite add (Light), "
            "swipe-delete (Medium), level-up / quest claim / achievement claim (Success notification)."
        ),
        xp_impact   = "—",
        thesis_tag  = "UX Design",
    )
    pm.add_feature(
        num         = 44,
        name        = "Streak system + badge",
        category    = "Gamification",
        session     = "S15",
        files       = "utils/streakSystem.ts (new), UserLevelCard.tsx, user.tsx",
        description = (
            "computeStreak(trips) counts consecutive days ending today or yesterday. "
            "Shown as 🔥 Xd badge in UserLevelCard XP footer row. Hidden at 0."
        ),
        xp_impact   = "—",
        thesis_tag  = "Gamification",
    )

    # ═══════════════════════════════════════════════════════════════════════════
    # SESSION 16 — 2026-03-16
    # Streak milestone rewards + Log Again feature
    # ═══════════════════════════════════════════════════════════════════════════

    # ── S16-1. DASHBOARD KPIs ─────────────────────────────────────────────────
    pm.update_dashboard_kpis(
        sessions  = 16,
        features  = 46,   # 44 (S15) + 2 new (streak milestones, log again)
        backlog   = 10,   # closed #47 (log again); streak milestones was new+done; 10 open remain
        tech_debt = 0,
        blockers  = 5,    # PP1-PP5 unchanged
    )

    # ── S16-2. VELOCITY ROW ───────────────────────────────────────────────────
    pm.add_velocity_row(
        session    = "S16",
        date       = "Mar 16",
        focus      = "Streak milestone XP rewards + Log Again (replaces #47 recurring trip)",
        features_n = 2,
        category   = "Gamification / UX",
    )

    # ── S16-3. SESSION DIARY ROW ──────────────────────────────────────────────
    pm.add_session_row(
        session      = "S16",
        date         = "2026-03-16",
        scope        = "Streak milestone rewards · Log Again button · QuickAdd prefill",
        what         = (
            "Streak milestones: STREAK_MILESTONES array (3/7/14/30/60/100 days → 50/150/300/600/1000/2000 XP) "
            "and getNewMilestoneXP(streak, claimed) added to utils/streakSystem.ts. "
            "claimedMilestones state in user.tsx loaded on init/focus from @claimedStreakMilestones. "
            "After each trip add, checkStreakMilestones fires; if a milestone is newly hit, "
            "bonus XP awarded 1.8s after trip toast (haptic + level-up if triggered). "
            "Log Again: onLogAgain prop added to TripDetailModal; view mode shows Close + Log Again "
            "two-button row (green, replay icon). "
            "QuickAddTripModal gains optional prefill prop (PrefillData type) with useEffect([visible]) "
            "that populates all fields when modal opens; date resets to now. "
            "handleLogAgain in user.tsx: closes detail modal → sets prefillData state → opens QuickAdd. "
            "Files: utils/streakSystem.ts, components/ui/TripDetailModal.tsx, "
            "components/ui/QuickAddTripModal.tsx, app/(tabs)/user.tsx (Write)"
        ),
        why          = (
            "Milestone rewards give the streak system progression depth beyond a simple counter — "
            "they reward users who maintain long streaks and make the streak feel meaningful. "
            "Log Again replaces the vague #47 'recurring trip' item: favorites already handle the "
            "high-frequency repeat case; Log Again fills the gap for any past trip regardless of frequency."
        ),
        problems     = (
            "TripDetailModal Edit tool failed on first attempt (indentation mismatch); "
            "resolved by re-reading exact whitespace before editing. "
            "Milestone XP toast coordination: two toasts in quick succession (trip XP + milestone XP) "
            "resolved by delaying milestone toast by 1.8s."
        ),
        alternatives = (
            "Milestone XP could be summed with trip XP and shown in one toast — rejected because "
            "the delay makes the milestone feel like a separate reward rather than noise. "
            "Log Again could have re-logged without opening QuickAdd — rejected because user "
            "should be able to adjust the date/time before confirming."
        ),
        decisions    = (
            "Milestone awards are one-time (persisted to @claimedStreakMilestones) — "
            "a streak breaking and restarting does NOT re-award. "
            "prefill date always resets to now (not the original trip date) — "
            "the intent is to log a new trip today, not to duplicate a past date."
        ),
        features_n   = 2,
        effort       = "~3h",
        thesis_tag   = "Gamification / Engagement",
    )

    # ── S16-4. MARK BACKLOG ITEMS DONE ────────────────────────────────────────
    pm.mark_backlog_done(
        "47", "S16",
        "Replaced by Log Again button in TripDetailModal + QuickAdd prefill prop"
    )

    # ── S16-5. FEATURE REGISTRY ───────────────────────────────────────────────
    pm.add_feature(
        num         = 45,
        name        = "Streak milestone XP rewards",
        category    = "Gamification",
        session     = "S16",
        files       = "utils/streakSystem.ts, app/(tabs)/user.tsx",
        description = (
            "STREAK_MILESTONES: 3/7/14/30/60/100 days → 50/150/300/600/1000/2000 XP. "
            "One-time auto-award on milestone hit. Delayed XP toast (1.8s) + haptic. "
            "Claimed set persisted to @claimedStreakMilestones."
        ),
        xp_impact   = "Yes",
        thesis_tag  = "Gamification",
    )
    pm.add_feature(
        num         = 46,
        name        = "Log Again + QuickAdd prefill",
        category    = "UX",
        session     = "S16",
        files       = "TripDetailModal.tsx, QuickAddTripModal.tsx, app/(tabs)/user.tsx",
        description = (
            "TripDetailModal view mode: Close + Log Again two-button row (green, replay icon). "
            "QuickAddTripModal prefill prop: all fields pre-populated from past trip, date resets to now."
        ),
        xp_impact   = "—",
        thesis_tag  = "UX Design",
    )

    # ── 7. SAVE ───────────────────────────────────────────────────────────────
    pm.save()
    print("\n=== Done. Narrative columns (Reflection, Theory Reference) "
          "left as placeholders — enrich via web interface. ===\n")
